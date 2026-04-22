#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Converte todas as abas de uma planilha Excel (.xlsx) em arquivos CSV separados.

Uso:
    python xlsx_para_csv_abas.py arquivo.xlsx

Opcional:
    python xlsx_para_csv_abas.py arquivo.xlsx --saida pasta_destino

Requisitos:
    pip install openpyxl
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


def limpar_nome_arquivo(nome: str) -> str:
    nome = nome.strip()
    nome = re.sub(r'[\\/:*?"<>|]+', "_", nome)
    nome = re.sub(r"\s+", " ", nome)
    return nome or "aba"


def converter_planilha(xlsx_path: Path, pasta_saida: Path) -> list[Path]:
    wb = load_workbook(xlsx_path, data_only=True)
    pasta_saida.mkdir(parents=True, exist_ok=True)

    arquivos_gerados: list[Path] = []

    for indice, ws in enumerate(wb.worksheets, start=1):
        nome_limpo = limpar_nome_arquivo(ws.title)
        csv_path = pasta_saida / f"{indice:02d}_{nome_limpo}.csv"

        with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if cell is None else cell for cell in row])

        arquivos_gerados.append(csv_path)

    return arquivos_gerados


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Converte todas as abas de um arquivo Excel em CSVs separados."
    )
    parser.add_argument("arquivo", help="Caminho do arquivo .xlsx")
    parser.add_argument(
        "--saida",
        help="Pasta de saída dos CSVs. Se omitido, será criada uma pasta ao lado do arquivo.",
        default=None,
    )

    args = parser.parse_args()

    xlsx_path = Path(args.arquivo).expanduser().resolve()

    if not xlsx_path.exists():
        print(f"Erro: arquivo não encontrado: {xlsx_path}")
        return 1

    if xlsx_path.suffix.lower() != ".xlsx":
        print("Erro: este script espera um arquivo .xlsx")
        return 1

    if args.saida:
        pasta_saida = Path(args.saida).expanduser().resolve()
    else:
        pasta_saida = xlsx_path.parent / f"{xlsx_path.stem}_csv"

    try:
        arquivos = converter_planilha(xlsx_path, pasta_saida)
    except Exception as e:
        print(f"Erro ao converter a planilha: {e}")
        return 1

    print("Conversão concluída.")
    print(f"Pasta de saída: {pasta_saida}")
    print("Arquivos gerados:")
    for arq in arquivos:
        print(f"- {arq.name}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
