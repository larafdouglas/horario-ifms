#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import csv
import re
import sys
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk

from openpyxl import load_workbook


def limpar_nome_arquivo(nome: str) -> str:
    nome = nome.strip()
    nome = re.sub(r'[\\/:*?"<>|]+', "_", nome)
    nome = re.sub(r"\s+", " ", nome)
    return nome or "aba"


def converter_planilha(xlsx_path: Path, pasta_saida: Path) -> list[Path]:
    wb = load_workbook(xlsx_path, data_only=True)
    pasta_saida.mkdir(parents=True, exist_ok=True)

    arquivos_gerados: list[Path] = []

    for indice, ws in enumerate(wb.worksheets, start=1):
        nome_limpo = limpar_nome_arquivo(ws.title)
        csv_path = pasta_saida / f"{indice:02d}_{nome_limpo}.csv"

        with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if cell is None else cell for cell in row])

        arquivos_gerados.append(csv_path)

    return arquivos_gerados


class AppEtapas:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Conversor por Etapas")
        self.root.geometry("820x560")
        self.root.minsize(760, 500)

        self.arquivo_var = tk.StringVar()
        self.saida_var = tk.StringVar()

        self.xlsx_path: Path | None = None
        self.pasta_saida: Path | None = None
        self.arquivos_gerados: list[Path] = []

        self.etapa1_concluida = False
        self.etapa2_concluida = False

        self.criar_interface()

    def criar_interface(self):
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=10)

        self.tab1 = tk.Frame(self.notebook)
        self.tab2 = tk.Frame(self.notebook)
        self.tab3 = tk.Frame(self.notebook)

        self.notebook.add(self.tab1, text="Etapa 1 - Gerar CSV")
        self.notebook.add(self.tab2, text="Etapa 2 - Próxima etapa")
        self.notebook.add(self.tab3, text="Etapa 3 - Finalização")

        # Travar abas seguintes no início
        self.notebook.tab(1, state="disabled")
        self.notebook.tab(2, state="disabled")

        self.montar_tab1()
        self.montar_tab2()
        self.montar_tab3()

    def montar_tab1(self):
        frame = tk.Frame(self.tab1, padx=12, pady=12)
        frame.pack(fill="both", expand=True)

        tk.Label(
            frame,
            text="Etapa 1 — Selecionar planilha e gerar CSVs",
            font=("Arial", 14, "bold")
        ).pack(anchor="w", pady=(0, 12))

        linha1 = tk.Frame(frame)
        linha1.pack(fill="x", pady=6)

        tk.Label(linha1, text="Arquivo .xlsx:", width=14, anchor="w").pack(side="left")
        tk.Entry(linha1, textvariable=self.arquivo_var).pack(
            side="left", fill="x", expand=True, padx=6
        )
        tk.Button(linha1, text="Procurar", command=self.escolher_arquivo).pack(side="left")

        linha2 = tk.Frame(frame)
        linha2.pack(fill="x", pady=6)

        tk.Label(linha2, text="Pasta de saída:", width=14, anchor="w").pack(side="left")
        tk.Entry(linha2, textvariable=self.saida_var).pack(
            side="left", fill="x", expand=True, padx=6
        )
        tk.Button(linha2, text="Escolher", command=self.escolher_saida).pack(side="left")

        tk.Label(
            frame,
            text="Se a pasta de saída ficar vazia, será criada automaticamente ao lado do arquivo.",
            fg="gray30"
        ).pack(anchor="w", pady=(4, 10))

        botoes = tk.Frame(frame)
        botoes.pack(fill="x", pady=(0, 10))

        tk.Button(
            botoes,
            text="Executar Etapa 1",
            command=self.executar_etapa1,
            width=18,
            height=2
        ).pack(side="left")

        tk.Button(
            botoes,
            text="Limpar",
            command=self.limpar_etapa1,
            width=18,
            height=2
        ).pack(side="left", padx=8)

        tk.Label(frame, text="Log da Etapa 1:").pack(anchor="w")
        self.log1 = scrolledtext.ScrolledText(frame, wrap="word", height=18)
        self.log1.pack(fill="both", expand=True)

    def montar_tab2(self):
        frame = tk.Frame(self.tab2, padx=12, pady=12)
        frame.pack(fill="both", expand=True)

        tk.Label(
            frame,
            text="Etapa 2 — Processamento seguinte",
            font=("Arial", 14, "bold")
        ).pack(anchor="w", pady=(0, 12))

        self.info_etapa2 = tk.Label(
            frame,
            text="Esta etapa será liberada quando a Etapa 1 for concluída.",
            justify="left"
        )
        self.info_etapa2.pack(anchor="w", pady=(0, 10))

        tk.Button(
            frame,
            text="Concluir Etapa 2",
            command=self.executar_etapa2,
            width=18,
            height=2
        ).pack(anchor="w", pady=(0, 10))

        tk.Label(frame, text="Log da Etapa 2:").pack(anchor="w")
        self.log2 = scrolledtext.ScrolledText(frame, wrap="word", height=18)
        self.log2.pack(fill="both", expand=True)

    def montar_tab3(self):
        frame = tk.Frame(self.tab3, padx=12, pady=12)
        frame.pack(fill="both", expand=True)

        tk.Label(
            frame,
            text="Etapa 3 — Finalização",
            font=("Arial", 14, "bold")
        ).pack(anchor="w", pady=(0, 12))

        self.info_etapa3 = tk.Label(
            frame,
            text="Esta etapa será liberada quando a Etapa 2 for concluída.",
            justify="left"
        )
        self.info_etapa3.pack(anchor="w", pady=(0, 10))

        tk.Label(frame, text="Log da Etapa 3:").pack(anchor="w")
        self.log3 = scrolledtext.ScrolledText(frame, wrap="word", height=18)
        self.log3.pack(fill="both", expand=True)

    def escrever_log(self, widget, texto: str):
        widget.insert(tk.END, texto + "\n")
        widget.see(tk.END)

    def escolher_arquivo(self):
        caminho = filedialog.askopenfilename(
            title="Selecione o arquivo Excel",
            filetypes=[("Arquivos Excel", "*.xlsx")]
        )
        if caminho:
            self.arquivo_var.set(caminho)

    def escolher_saida(self):
        pasta = filedialog.askdirectory(title="Selecione a pasta de saída")
        if pasta:
            self.saida_var.set(pasta)

    def limpar_etapa1(self):
        self.arquivo_var.set("")
        self.saida_var.set("")
        self.log1.delete("1.0", tk.END)

    def executar_etapa1(self):
        self.log1.delete("1.0", tk.END)

        arquivo = self.arquivo_var.get().strip()
        saida = self.saida_var.get().strip()

        if not arquivo:
            messagebox.showwarning("Aviso", "Selecione um arquivo .xlsx.")
            return

        xlsx_path = Path(arquivo).expanduser().resolve()

        if not xlsx_path.exists():
            messagebox.showerror("Erro", f"Arquivo não encontrado:\n{xlsx_path}")
            return

        if xlsx_path.suffix.lower() != ".xlsx":
            messagebox.showerror("Erro", "Este programa espera um arquivo .xlsx")
            return

        if saida:
            pasta_saida = Path(saida).expanduser().resolve()
        else:
            pasta_saida = xlsx_path.parent / f"{xlsx_path.stem}_csv"

        try:
            arquivos = converter_planilha(xlsx_path, pasta_saida)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao converter a planilha:\n{e}")
            self.escrever_log(self.log1, f"Erro: {e}")
            return

        self.xlsx_path = xlsx_path
        self.pasta_saida = pasta_saida
        self.arquivos_gerados = arquivos
        self.etapa1_concluida = True

        self.escrever_log(self.log1, "Etapa 1 concluída com sucesso.")
        self.escrever_log(self.log1, f"Pasta de saída: {pasta_saida}")
        self.escrever_log(self.log1, "Arquivos gerados:")
        for arq in arquivos:
            self.escrever_log(self.log1, f"- {arq.name}")

        # Libera a etapa 2
        self.notebook.tab(1, state="normal")
        self.info_etapa2.config(
            text=(
                "Etapa 2 liberada.\n\n"
                f"Arquivo original: {self.xlsx_path}\n"
                f"Pasta dos CSVs: {self.pasta_saida}\n"
                f"Total de arquivos gerados: {len(self.arquivos_gerados)}"
            )
        )

        messagebox.showinfo("Sucesso", "Etapa 1 concluída. A Etapa 2 foi liberada.")
        self.notebook.select(1)

    def executar_etapa2(self):
        if not self.etapa1_concluida:
            messagebox.showwarning("Aviso", "Conclua a Etapa 1 antes.")
            return

        self.log2.delete("1.0", tk.END)
        self.escrever_log(self.log2, "Etapa 2 concluída com sucesso.")
        self.escrever_log(self.log2, "Aqui você vai colocar a lógica real da etapa 2.")

        self.etapa2_concluida = True

        # Libera a etapa 3
        self.notebook.tab(2, state="normal")
        self.info_etapa3.config(
            text="Etapa 3 liberada. Aqui você pode colocar a finalização do processo."
        )

        messagebox.showinfo("Sucesso", "Etapa 2 concluída. A Etapa 3 foi liberada.")
        self.notebook.select(2)


def main() -> int:
    root = tk.Tk()
    app = AppEtapas(root)
    root.mainloop()
    return 0


if __name__ == "__main__":
    sys.exit(main())